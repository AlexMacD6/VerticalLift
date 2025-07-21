import os
from fastapi import FastAPI, UploadFile, Form, Depends, HTTPException, status, Path, Query
from fastapi.middleware.cors import CORSMiddleware
import pandas as pd, optimiser, io, json
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker, Session
from models import Base, TrayConfig, Inventory
from dotenv import load_dotenv
import traceback
import numpy as np

load_dotenv()  # Loads .env file from project root

DATABASE_URL = os.getenv("DATABASE_URL")

# Initialize database connection lazily
engine = None
SessionLocal = None

def get_engine():
    global engine
    if engine is None:
        if not DATABASE_URL:
            raise ValueError("DATABASE_URL environment variable is not set")
        engine = create_engine(DATABASE_URL, pool_pre_ping=True)
        Base.metadata.create_all(bind=engine)
    return engine

def get_session_local():
    global SessionLocal
    if SessionLocal is None:
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=get_engine())
    return SessionLocal

def get_db():
    db = get_session_local()()
    try:
        yield db
    finally:
        db.close()

app = FastAPI(title="Tray Optimizer MVP")

# Add CORS middleware (allow all origins for development)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Allow all origins for dev
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/")
def read_root():
    return {"message": "Tray Optimizer MVP API is running"}

@app.get("/health")
def health_check():
    return {"status": "healthy", "database_url_set": bool(DATABASE_URL)}

def convert_numpy(obj):
    if isinstance(obj, dict):
        return {k: convert_numpy(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [convert_numpy(i) for i in obj]
    elif hasattr(obj, 'item') and callable(obj.item):
        try:
            return obj.item()
        except Exception:
            return str(obj)
    elif isinstance(obj, float) and (obj != obj):  # Check for NaN
        return 0.0
    elif isinstance(obj, float) and (obj == float('inf') or obj == float('-inf')):
        return 0.0
    else:
        return obj

@app.get("/tray-configs")
def get_tray_configs(db: Session = Depends(get_db)):
    return db.query(TrayConfig).all()

@app.post("/tray-configs")
def create_tray_config(config: dict, db: Session = Depends(get_db)):
    print("[POST /tray-configs] Incoming config:", config)
    try:
        tray = TrayConfig(**config)
        db.add(tray)
        db.commit()
        db.refresh(tray)
        print("[POST /tray-configs] Saved tray config with id:", tray.id)
        return tray
    except Exception as e:
        db.rollback()
        print("[POST /tray-configs] Error:", str(e))
        traceback.print_exc()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Failed to save tray config: {str(e)}"
        )

@app.put("/tray-configs/{config_id}")
def update_tray_config(config_id: int = Path(...), config: dict = None, db: Session = Depends(get_db)):
    tray = db.query(TrayConfig).filter(TrayConfig.id == config_id).first()
    if not tray:
        raise HTTPException(status_code=404, detail="Tray configuration not found")
    for key, value in config.items():
        if hasattr(tray, key):
            setattr(tray, key, value)
    db.commit()
    db.refresh(tray)
    return tray

@app.post("/import-inventory")
async def import_inventory(file: UploadFile, inventory_list_id: str = Form(None), db: Session = Depends(get_db)):
    """Import inventory data from CSV or Excel into database. Optionally associate with an inventory list."""
    print(f"[POST /import-inventory] Processing file: {file.filename}")
    try:
        file_content = await file.read()
        df = None
        # Determine file type and read accordingly
        if file.filename.lower().endswith('.csv'):
            try:
                df = pd.read_csv(io.BytesIO(file_content), header=2)  # Try row 3 as header
                if df.columns[0] not in ['SKU', 'sku_id']:
                    # If not the expected header, try first row
                    df = pd.read_csv(io.BytesIO(file_content), header=0)
            except Exception as e:
                print(f"[POST /import-inventory] CSV read error: {e}, trying header=0 fallback")
                df = pd.read_csv(io.BytesIO(file_content), header=0)
        elif file.filename.lower().endswith(('.xlsx', '.xls')):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
            # Ignore 'Definitions' sheet for validation
            sheet_names = [name for name in wb.sheetnames if name != 'Definitions']
            print(f"[POST /import-inventory] Excel sheets (excluding Definitions): {sheet_names}")
            # Annual template: single sheet, daily template: two sheets
            if len(sheet_names) == 1 and sheet_names[0] in ["SKU Master", "Inventory Template"]:
                ws = wb[sheet_names[0]]
                headers = [cell.value for cell in next(ws.iter_rows(min_row=3, max_row=3))]
                required = [
                    'SKU', 'Product Name', 'Length (in)', 'Width (in)', 'Height (in)',
                    'Weight (lb)', 'In Stock', 'Annual Sales'
                ]
                if not all(h in headers for h in required):
                    raise HTTPException(status_code=400, detail=f"Missing required columns in SKU Master: {required}")
                # Read data into DataFrame
                df = pd.DataFrame(ws.iter_rows(min_row=4, values_only=True), columns=headers)
            elif len(sheet_names) == 2 and set(sheet_names) == {"SKU Master", "Daily Sales"}:
                ws_sku = wb["SKU Master"]
                ws_daily = wb["Daily Sales"]
                sku_headers = [cell.value for cell in next(ws_sku.iter_rows(min_row=3, max_row=3))]
                daily_headers = [cell.value for cell in next(ws_daily.iter_rows(min_row=3, max_row=3))]
                sku_required = [
                    'SKU', 'Product Name', 'Length (in)', 'Width (in)', 'Height (in)',
                    'Weight (lb)', 'In Stock', 'Annual Sales'
                ]
                daily_required = ['Date', 'SKU', 'Units Sold']
                if not all(h in sku_headers for h in sku_required):
                    raise HTTPException(status_code=400, detail=f"Missing required columns in SKU Master: {sku_required}")
                if not all(h in daily_headers for h in daily_required):
                    raise HTTPException(status_code=400, detail=f"Missing required columns in Daily Sales: {daily_required}")
                # Read data into DataFrames
                df_sku = pd.DataFrame(ws_sku.iter_rows(min_row=4, values_only=True), columns=sku_headers)
                df_daily = pd.DataFrame(ws_daily.iter_rows(min_row=4, values_only=True), columns=daily_headers)
                # Use SKU Master for inventory import
                df = df_sku
                # Store daily sales data for later processing
                daily_sales_df = df_daily
            else:
                raise HTTPException(status_code=400, detail="Excel file must have either one sheet named 'SKU Master' or two sheets named 'SKU Master' and 'Daily Sales'. (Other sheets like 'Definitions' are ignored.)")
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Unsupported file format. Please upload a CSV or Excel file."
            )
        print(f"[POST /import-inventory] Loaded {len(df)} rows, columns: {df.columns.tolist()}")
        # Map new column names to database column names
        column_mapping = {
            'SKU': 'sku_id',
            'Product Name': 'description',
            'Length (in)': 'length_in',
            'Width (in)': 'width_in',
            'Height (in)': 'height_in',
            'Weight (lb)': 'weight_lb',
            'In Stock': 'on_hand_units',
            'Annual Sales': 'annual_units_sold',
        }
        # Rename columns if they match the new template format
        df_columns = df.columns.tolist()
        if any(col in column_mapping for col in df_columns):
            df = df.rename(columns=column_mapping)
        # Drop empty rows (all NaN)
        df = df.dropna(how='all')
        
        # Clear existing inventory for this specific list only
        if inventory_list_id:
            db.query(Inventory).filter(Inventory.inventory_list_id == inventory_list_id).delete()
        else:
            # If no inventory_list_id provided, clear all inventory (fallback behavior)
            db.query(Inventory).delete()
            
        # Import new data
        for _, row in df.iterrows():
            # Skip rows missing required fields
            if pd.isna(row.get('sku_id')):
                continue
                
            # Ensure inventory_list_id is provided
            if not inventory_list_id:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="inventory_list_id is required for importing inventory"
                )
                
            inventory_item = Inventory(
                sku_id=row['sku_id'],
                description=row['description'],
                length_in=row['length_in'],
                width_in=row['width_in'],
                height_in=row['height_in'],
                weight_lb=row['weight_lb'],
                on_hand_units=row['on_hand_units'],
                annual_units_sold=row['annual_units_sold'],
                daily_picks=None,
                demand_std_dev=None,
                inventory_list_id=inventory_list_id
            )
            db.add(inventory_item)
        
        # Process daily sales data if available
        daily_sales_count = 0
        if 'daily_sales_df' in locals() and not daily_sales_df.empty and inventory_list_id:
            from models import DailySales
            # Clear existing daily sales for this list
            db.query(DailySales).filter(DailySales.inventory_list_id == inventory_list_id).delete()
            
            # Import daily sales data
            for _, row in daily_sales_df.iterrows():
                if pd.isna(row.get('Date')) or pd.isna(row.get('SKU')) or pd.isna(row.get('Units Sold')):
                    continue
                    
                try:
                    # Parse date - handle multiple formats
                    date_str = str(row['Date'])
                    if '/' in date_str:
                        # Handle MM/DD/YYYY format
                        date_obj = pd.to_datetime(date_str, format='%m/%d/%Y')
                    else:
                        # Try other formats
                        date_obj = pd.to_datetime(date_str)
                        
                    daily_sales_item = DailySales(
                        date=date_obj,
                        sku_id=str(row['SKU']),
                        units_sold=int(row['Units Sold']),
                        inventory_list_id=inventory_list_id
                    )
                    db.add(daily_sales_item)
                    daily_sales_count += 1
                except Exception as e:
                    print(f"Error processing daily sales row: {row}, error: {e}")
                    continue
        
        db.commit()
        message = f"Successfully imported {len(df)} inventory items"
        if daily_sales_count > 0:
            message += f" and {daily_sales_count} daily sales records"
        print(f"[POST /import-inventory] {message}")
        return {"message": message}
    except Exception as e:
        db.rollback()
        print(f"[POST /import-inventory] Error: {str(e)}")
        traceback.print_exc()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Failed to import inventory: {str(e)}"
        )

@app.get("/inventory")
def get_inventory(inventory_list_id: str = Query(None), db: Session = Depends(get_db)):
    """Get all inventory items, or those for a specific inventory list."""
    query = db.query(Inventory)
    if inventory_list_id:
        query = query.filter(Inventory.inventory_list_id == inventory_list_id)
    return query.all()

@app.get("/daily-sales")
def get_daily_sales(inventory_list_id: str = Query(None), db: Session = Depends(get_db)):
    """Get daily sales data for a specific inventory list."""
    from models import DailySales
    query = db.query(DailySales)
    if inventory_list_id:
        query = query.filter(DailySales.inventory_list_id == inventory_list_id)
    return query.all()

@app.post("/inventory-lists/{list_id}/copy-inventory")
def copy_inventory_to_list(list_id: str, db: Session = Depends(get_db)):
    """Copy all current inventory items to the specified inventory list."""
    from models import Inventory
    items = db.query(Inventory).all()
    for item in items:
        new_item = Inventory(
            sku_id=item.sku_id,
            description=item.description,
            length_in=item.length_in,
            width_in=item.width_in,
            height_in=item.height_in,
            weight_lb=item.weight_lb,
            on_hand_units=item.on_hand_units,
            annual_units_sold=item.annual_units_sold,
            daily_picks=item.daily_picks,
            demand_std_dev=item.demand_std_dev,
            inventory_list_id=list_id
        )
        db.add(new_item)
    db.commit()
    return {"message": f"Copied {len(items)} inventory items to list {list_id}"}

@app.get("/inventory-lists")
def get_inventory_lists(db: Session = Depends(get_db)):
    from models import InventoryList
    return db.query(InventoryList).all()

@app.post("/inventory-lists")
def create_inventory_list(list_data: dict, db: Session = Depends(get_db)):
    from models import InventoryList
    try:
        inventory_list = InventoryList(name=list_data["name"])
        db.add(inventory_list)
        db.commit()
        db.refresh(inventory_list)
        return inventory_list
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Failed to create inventory list: {str(e)}"
        )

@app.post("/inventory-lists/{list_id}/update-on-shelf-units")
def update_on_shelf_units(list_id: str, on_shelf_data: dict, db: Session = Depends(get_db)):
    """Update on_shelf_units for inventory items based on analytics calculations."""
    try:
        updated_count = 0
        data_list = on_shelf_data.get("on_shelf_data", [])
        for item in data_list:
            inventory_item = db.query(Inventory).filter(
                Inventory.sku_id == item["sku_id"],
                Inventory.inventory_list_id == list_id
            ).first()
            
            if inventory_item:
                inventory_item.on_shelf_units = item["on_shelf_units"]
                updated_count += 1
        
        db.commit()
        return {"message": f"Updated on_shelf_units for {updated_count} items"}
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Failed to update on_shelf_units: {str(e)}"
        )

@app.post("/optimize")
async def optimize(
    tray_length_in: int = Form(156),
    tray_width_in: int  = Form(36),
    tray_depth_in: int  = Form(18),
    num_trays: int      = Form(20),
    weight_limit_lb: int= Form(2205),
    buffer_pct: float   = Form(0.95),
    model: str = Form("rectpack"),
    inventory_list_id: str = Form(None),
    db: Session = Depends(get_db)
):
    """Return tray plan JSON using database inventory and selected model."""
    try:
        # Get inventory from database
        query = db.query(Inventory)
        if inventory_list_id:
            query = query.filter(Inventory.inventory_list_id == inventory_list_id)
        inventory_items = query.all()
        
        if not inventory_items:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No inventory data found. Please import inventory first."
            )
        
        # Convert to DataFrame
        df_data = []
        for item in inventory_items:
            df_data.append({
                'sku_id': item.sku_id,
                'description': item.description,
                'length_in': item.length_in,
                'width_in': item.width_in,
                'height_in': item.height_in,
                'weight_lb': item.weight_lb,
                'on_hand_units': item.on_hand_units,
                'on_shelf_units': item.on_shelf_units,  # Include calculated on-shelf units
                'annual_units_sold': item.annual_units_sold,
                'daily_picks': item.daily_picks,
                'demand_std_dev': item.demand_std_dev
            })
        
        import pandas as pd
        df = pd.DataFrame(df_data)
        print(f"[POST /optimize] Optimizing {len(df)} inventory items with model: {model}")
        
        # Run optimization based on model
        if model in ["rectpack", "maximal-rectangles"]:
            plan = optimiser.optimise(
                df,
                model=model,
                tray_length_in=tray_length_in,
                tray_width_in=tray_width_in,
                tray_depth_in=tray_depth_in,
                num_trays=num_trays,
                weight_limit_lb=weight_limit_lb,
                buffer_pct=buffer_pct,
            )
            
            # Calculate appropriate KPIs
            kpis = optimiser.calculate_kpis(
                plan,
                tray_length_in,
                tray_width_in,
                tray_depth_in,
                weight_limit_lb
            )
        else:
            raise HTTPException(status_code=400, detail=f"Unknown model: {model}")
        
        plan_records = plan.to_dict(orient="records")
        plan_records = convert_numpy(plan_records)
        kpis = convert_numpy(kpis)
        return {"plan": plan_records, "model": model, "kpis": kpis}
    except Exception as e:
        print(f"[POST /optimize] Error: {str(e)}")
        traceback.print_exc()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Optimization failed: {str(e)}"
        )

@app.post("/optimize-dividers")
async def optimize_dividers(
    tray_length_in: int = Form(156),
    tray_width_in: int  = Form(36),
    tray_depth_in: int  = Form(18),
    buffer_pct: float   = Form(0.95),
    model: str = Form("rectpack"),
    inventory_list_id: str = Form(None),
    db: Session = Depends(get_db)
):
    """Optimize divider sizes for each SKU using rectpack algorithm."""
    try:
        # Get inventory from database
        query = db.query(Inventory)
        if inventory_list_id:
            query = query.filter(Inventory.inventory_list_id == inventory_list_id)
        inventory_items = query.all()
        
        if not inventory_items:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No inventory data found. Please import inventory first."
            )
        
        # Convert to DataFrame
        df_data = []
        for item in inventory_items:
            df_data.append({
                'sku_id': item.sku_id,
                'description': item.description,
                'length_in': item.length_in,
                'width_in': item.width_in,
                'height_in': item.height_in,
                'weight_lb': item.weight_lb,
                'on_hand_units': item.on_hand_units,
                'on_shelf_units': item.on_shelf_units,  # Include calculated on-shelf units
                'annual_units_sold': item.annual_units_sold,
                'daily_picks': item.daily_picks,
                'demand_std_dev': item.demand_std_dev
            })
        
        import pandas as pd
        df = pd.DataFrame(df_data)
        print(f"[POST /optimize-dividers] Optimizing dividers for {len(df)} SKUs with model: {model}")
        
        # Run divider optimization
        if model in ["rectpack", "maximal-rectangles"]:
            result = optimiser.optimise(
                df,
                model=model,
                tray_length_in=tray_length_in,
                tray_width_in=tray_width_in,
                tray_depth_in=tray_depth_in,
                buffer_pct=buffer_pct,
                inventory_list_id=inventory_list_id,
            )
            
            # Calculate divider-specific KPIs
            kpis = optimiser.calculate_divider_kpis(
                result,
                tray_length_in,
                tray_width_in,
                tray_depth_in,
                buffer_pct
            )
        else:
            raise HTTPException(status_code=400, detail=f"Divider optimization supports rectpack models only. Got: {model}")
        
        result_records = result.to_dict(orient="records")
        result_records = convert_numpy(result_records)
        kpis = convert_numpy(kpis)
        
        # Debug: Print what we're returning
        print(f"[POST /optimize-dividers] Returning {len(result_records)} divider records")
        print(f"[POST /optimize-dividers] First record keys: {list(result_records[0].keys()) if result_records else 'No records'}")
        print(f"[POST /optimize-dividers] KPI keys: {list(kpis.keys())}")
        
        # Include tray layout data if available (for rectpack model)
        tray_layouts = result.attrs.get('tray_layouts', [])
        print(f"[POST /optimize-dividers] Tray layouts available: {len(tray_layouts)}")
        print(f"[POST /optimize-dividers] Tray layouts data: {tray_layouts}")
        
        return {
            "dividers": result_records,
            "kpis": kpis,
            "model": model,
            "trayLayouts": tray_layouts
        }
    except Exception as e:
        print(f"[POST /optimize-dividers] Error: {str(e)}")
        traceback.print_exc()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Divider optimization failed: {str(e)}"
        ) 